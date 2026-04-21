"""Parser del Excel 'Comisiones Gerentes Productos' (una hoja por GP)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.models import Antiguedad, Company, PersonaInput, Role

from ._utils import read_sheet_as_dicts, safe_str, to_float


def parse_gerentes_productos(path: str | Path) -> list[PersonaInput]:
    path = str(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()

    out: list[PersonaInput] = []
    for sn in sheets:
        rows = read_sheet_as_dicts(path, sn)
        if not rows:
            continue
        # Agregamos todo el resumen en un solo registro (una hoja = un GP)
        cantidad = len(rows)
        monto = sum(to_float(r.get("ValorBien")) for r in rows)
        r0 = rows[0]
        out.append(
            PersonaInput(
                cedula=safe_str(r0.get("CodAsesor")) or sn,
                nombre=sn.strip(),
                codigo=r0.get("CodAsesor"),
                company=Company.AUTO,
                role=Role.GERENTE_PRODUCTO,
                structure_id="gp_auto",
                antiguedad=Antiguedad.ANTIGUO,
                cantidad_contratos=cantidad,
                monto_total_contratos=monto,
                monto_mm=monto / 1_000_000,
                porcentaje_persistencia=to_float(r0.get("PorcentajePersistencia")),
                porcentaje_segundo_pago=to_float(r0.get("PorcentajeSegundoPagoRegionCompleta")),
                sistema_valor_salario=to_float(r0.get("Salario")),
                sistema_valor_bonificacion=to_float(r0.get("Bono")),
                sistema_monto_comision=sum(to_float(r.get("ValorComision")) for r in rows),
            )
        )
    return out
