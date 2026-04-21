"""Parsers de los 5 tipos de Excel de entrada."""
from __future__ import annotations

from pathlib import Path

from app.models import PersonaInput

from .asesores_fonbienes import parse_asesores_fonbienes
from .asesores_serven import parse_asesores_serven
from .gerentes_equipos import parse_gerentes_equipos
from .gerentes_productos import parse_gerentes_productos
from .gerentes_regionales import parse_gerentes_regionales


def detect_and_parse(path: Path) -> list[PersonaInput]:
    """Detecta el tipo de archivo por nombre de hojas y llama al parser correcto."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    joined = " | ".join(sheets).lower()

    if "resumen asesores fonbienes" in joined:
        return parse_asesores_fonbienes(path)
    if "resumen asesores serven" in joined:
        return parse_asesores_serven(path)
    if "segundo pago gerentes equipo" in joined or "persistencia gerentes equipo" in joined:
        return parse_gerentes_equipos(path)

    # GP y GR: una hoja por persona. Distinguimos por presencia de la columna "RegionMontoMinimo"
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheets[0]]
        header = next(ws.iter_rows(values_only=True), ()) or ()
    finally:
        wb.close()
    header_norm = [str(h or "").lower().replace(" ", "") for h in header]
    if any("promediomonto" in h or "contratoregion" in h for h in header_norm):
        if any("regionmontominimo" in h for h in header_norm):
            return parse_gerentes_productos(path)
        return parse_gerentes_regionales(path)

    raise ValueError(f"No se pudo detectar el tipo de Excel: {Path(path).name} (hojas: {sheets})")


__all__ = [
    "detect_and_parse",
    "parse_asesores_fonbienes",
    "parse_asesores_serven",
    "parse_gerentes_equipos",
    "parse_gerentes_productos",
    "parse_gerentes_regionales",
]
