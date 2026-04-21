"""Utilidades compartidas por los parsers."""
from __future__ import annotations

from typing import Any

from openpyxl import load_workbook


def read_sheet_as_dicts(path: str, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
    except KeyError:
        wb.close()
        return []
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None or v == "" for v in row):
            continue
        out.append(dict(zip(header, row, strict=False)))
    return out


def find_sheet(path: str, keyword: str) -> str | None:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        kw = keyword.lower()
        for sn in wb.sheetnames:
            if kw in sn.lower():
                return sn
    finally:
        wb.close()
    return None


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if isinstance(v, str):
            v = v.strip().replace(",", ".").replace("$", "").replace(" ", "")
        return float(v)
    except (TypeError, ValueError):
        return default


def to_int(v: Any, default: int = 0) -> int:
    return int(to_float(v, default))


def safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()
