"""Genera el Excel consolidado de nómina."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import ComputedCommission

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DISCREP_FILL = PatternFill("solid", fgColor="FDECEA")

COLUMNS = [
    ("Cédula", "cedula"),
    ("Nombre", "nombre"),
    ("Compañía", "company"),
    ("Rol", "role"),
    ("Estructura", "structure_id"),
    ("# Contratos", "cantidad_contratos"),
    ("Persistencia", "porcentaje_persistencia"),
    ("Segundo Pago %", "porcentaje_segundo_pago"),
    ("Monto base", "monto_base_comisionable"),
    ("% Comisión", "porcentaje_comision"),
    ("Factor variable", "factor_variable_persistencia"),
    ("Factor Segundo Pago", "factor_segundo_pago"),
    ("Comisión base", "valor_comision_base"),
    ("Comisión final", "valor_comision_final"),
    ("Garantizado", "valor_garantizado"),
    ("Bono", "valor_bono"),
    ("Bono final", "valor_bono_final"),
    ("Ajuste manual", "ajuste_manual"),
    ("Motivo ajuste", "motivo_ajuste"),
    ("TOTAL a pagar", "valor_total_a_pagar"),
    ("Discrepancia", "discrepancia"),
    ("Notas", "notas"),
]


def _fmt(value, field: str):
    if field in ("porcentaje_persistencia", "porcentaje_segundo_pago", "porcentaje_comision",
                 "factor_variable_persistencia", "factor_segundo_pago"):
        return value if value is None else float(value)
    if field == "notas" and isinstance(value, list):
        return " | ".join(value)
    return value


def build_consolidated_excel(resultados: list[ComputedCommission]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    for col_idx, (title, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, res in enumerate(resultados, start=2):
        d = res.model_dump()
        for col_idx, (_, field) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_fmt(d.get(field), field))
            if res.discrepancia:
                cell.fill = DISCREP_FILL

    # Anchos y formatos
    for col_idx, (_, field) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if field in ("valor_comision_base", "valor_comision_final", "valor_garantizado",
                     "valor_bono", "valor_bono_final", "valor_total_a_pagar",
                     "monto_base_comisionable", "ajuste_manual"):
            for r in range(2, len(resultados) + 2):
                ws[f"{letter}{r}"].number_format = "$#,##0"
            ws.column_dimensions[letter].width = 16
        elif field in ("porcentaje_persistencia", "porcentaje_segundo_pago", "porcentaje_comision",
                       "factor_variable_persistencia", "factor_segundo_pago"):
            for r in range(2, len(resultados) + 2):
                ws[f"{letter}{r}"].number_format = "0.00%"
            ws.column_dimensions[letter].width = 12
        elif field == "nombre":
            ws.column_dimensions[letter].width = 34
        elif field == "notas":
            ws.column_dimensions[letter].width = 50
        else:
            ws.column_dimensions[letter].width = 14

    ws.freeze_panes = "A2"

    # Hoja de totales
    totals = wb.create_sheet("Totales")
    totals.append(["Métrica", "Valor"])
    total_pagar = sum(r.valor_total_a_pagar for r in resultados)
    totals.append(["Total a pagar", total_pagar])
    totals.append(["# Personas", len(resultados)])
    totals.append(["# Discrepancias", sum(1 for r in resultados if r.discrepancia)])
    totals["B2"].number_format = "$#,##0"
    totals.column_dimensions["A"].width = 28
    totals.column_dimensions["B"].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def save_consolidated_excel(resultados: list[ComputedCommission], path: Path) -> Path:
    content = build_consolidated_excel(resultados)
    path.write_bytes(content)
    return path
